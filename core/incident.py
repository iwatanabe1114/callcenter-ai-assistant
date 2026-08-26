"""インシデントテンプレートの読み込みと管理。

「インシデントテンプレ （202411~）」シートは横並びブロック構造：
- 行1: 注意書き（スキップ）
- 行2: ブロック名（通常便/おまとめ便_解約, 回数便解約, 継続応援成功, 注文キャンセル, その他）
- 行3: 各ブロックの列ヘッダー（タイトル, 内容, VOC, 注文状況, 継続応援結果, 解約希望理由…）
- 行4以降: テンプレートデータ

Claude APIは使用しない。テンプレート選択とフォーム入力のみ。
"""

from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

from .config import CACHE_DIR

INCIDENT_CACHE_PATH = CACHE_DIR / "incident_templates.json"


@dataclass
class IncidentTemplate:
    category: str
    title: str
    content: str
    voc: str = ""
    order_status: str = ""
    retention_result: str = ""
    cancel_reason: str = ""
    extras: dict[str, str] = field(default_factory=dict)


def parse_incident_rows(rows: list[list[str]]) -> list[IncidentTemplate]:
    """縦並びブロック構造のシートをパースしてテンプレート一覧を返す。

    構造:
        カテゴリ名の行（A列のみ値、B列以降は空）
        列ヘッダーの行（タイトル, 内容, VOC, ...）
        データ行
        空行
        次のカテゴリ名の行
        ...
    """
    if len(rows) < 3:
        return []

    templates: list[IncidentTemplate] = []
    current_category = ""
    col_map: dict[str, int] = {}

    for r, row in enumerate(rows):
        cells = [str(c).strip() for c in row]

        # 空行 → スキップ
        if not any(cells):
            continue

        # 行1は注意書き（スキップ）
        if r == 0:
            continue

        first = cells[0]
        rest_empty = all(c == "" for c in cells[1:8])

        # カテゴリ名の行: A列に値があり、B列以降がほぼ空、「タイトル」ではない
        if first and first != "タイトル" and rest_empty:
            # 「入電時の対応フロー」以降はテンプレートではないので終了
            if first == "入電時の対応フロー":
                break
            current_category = first
            continue

        # 列ヘッダー行: A列が「タイトル」
        if first == "タイトル":
            col_map = {}
            for i, c in enumerate(cells):
                col_name = c.split("\n")[0].strip()
                if col_name:
                    col_map[col_name] = i
            continue

        # データ行
        if not current_category or "タイトル" not in col_map:
            continue

        title_idx = col_map.get("タイトル")
        content_idx = col_map.get("内容")
        if title_idx is None or content_idx is None:
            continue

        def _cell(idx: int | None) -> str:
            if idx is None or idx >= len(cells):
                return ""
            return cells[idx]

        title = _cell(title_idx)
        content = _cell(content_idx)
        if not title and not content:
            continue

        voc = _cell(col_map.get("VOC"))
        order_status = ""
        retention_result = ""
        cancel_reason = ""
        extras: dict[str, str] = {}

        for col_name, col_idx in col_map.items():
            if col_name in ("タイトル", "内容", "VOC"):
                continue
            val = _cell(col_idx)
            if not val:
                continue
            if "注文状況" in col_name:
                order_status = val
            elif "継続応援結果" in col_name:
                retention_result = val
            elif "解約希望理由" in col_name:
                cancel_reason = val
            else:
                extras[col_name] = val

        templates.append(IncidentTemplate(
            category=current_category,
            title=title,
            content=content,
            voc=voc,
            order_status=order_status,
            retention_result=retention_result,
            cancel_reason=cancel_reason,
            extras=extras,
        ))

    return templates


# ── キャッシュ ──────────────────────────────────────────────

def save_cache(templates: list[IncidentTemplate]) -> Path:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    data = [asdict(t) for t in templates]
    INCIDENT_CACHE_PATH.write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return INCIDENT_CACHE_PATH


def load_cache() -> list[IncidentTemplate]:
    if not INCIDENT_CACHE_PATH.exists():
        return []
    data = json.loads(INCIDENT_CACHE_PATH.read_text(encoding="utf-8"))
    return [IncidentTemplate(**d) for d in data]


# ── ユーティリティ ──────────────────────────────────────────

def get_categories(templates: list[IncidentTemplate]) -> list[str]:
    """カテゴリを登場順で返す。"""
    seen: list[str] = []
    for t in templates:
        if t.category not in seen:
            seen.append(t.category)
    return seen


def fill_title(title: str, product_name: str) -> str:
    """タイトル内の「商品名」を実際の商品名に置換する。"""
    return title.replace("商品名", product_name)


def fill_content(content: str, product_name: str) -> str:
    """内容内の商品名プレースホルダーを置換する。"""
    result = content
    result = re.sub(r"【〇〇", f"【{product_name}", result)
    result = re.sub(r"【○○", f"【{product_name}", result)
    result = re.sub(r"^〇〇", product_name, result)
    result = re.sub(r"^○○", product_name, result)
    return result


# ── xlsx / gspread からの読み込み ───────────────────────────

def load_from_xlsx(path: Path, worksheet_name: str) -> list[IncidentTemplate]:
    """Excelファイルからインシデントテンプレートを読み込む。"""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if worksheet_name not in wb.sheetnames:
            return []
        ws = wb[worksheet_name]
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(c).strip() if c is not None else "" for c in row])
    finally:
        wb.close()

    return parse_incident_rows(rows)


def load_from_gspread(
    service_account_info: dict[str, Any],
    spreadsheet_id: str,
    worksheet_name: str,
) -> list[IncidentTemplate]:
    """Google スプレッドシートからインシデントテンプレートを読み込む。"""
    import gspread
    from google.oauth2.service_account import Credentials

    creds = Credentials.from_service_account_info(
        service_account_info,
        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
    )
    gc = gspread.authorize(creds)
    book = gc.open_by_key(spreadsheet_id)
    ws = book.worksheet(worksheet_name)
    rows = ws.get_all_values()
    return parse_incident_rows(rows)
