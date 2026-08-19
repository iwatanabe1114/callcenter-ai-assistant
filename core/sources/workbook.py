"""Excel（.xlsx）ファイルからの取得。

サービスアカウントを用意する前に、実データで動作確認するための経路です。
スプレッドシートを「ファイル」→「ダウンロード」→「Microsoft Excel (.xlsx)」で
書き出したものを読みます。

なぜこの方法なのか:
    公開URLからCSVで取る方法もありますが、Googleのgviz書き出しはシートによって
    複数行を1行に結合したり、2段ヘッダーの扱いが変わったりします。
    取得経路によってデータが変わると、動作確認の意味がなくなります。
    Excel書き出しはセルの並びがそのまま保たれるため、本番（gspread）と
    同じ行データが得られます。

読み取り部分（列の指定・日付での除外・縦持ち）は
sheets.py と同じ処理を使い回します。
"""

from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

from ..config import AppConfig
from . import sheets


class WorkbookError(RuntimeError):
    pass


def _cell_text(value: Any) -> str:
    """Excelのセルの値を、スプレッドシートの見た目に近い文字列にする。

    openpyxl は数値を float、日付を datetime で返すため、そのまま str() すると
    「500.0」「2026-08-06 00:00:00」のようになり、料金や日付が読みにくくなります。
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d")
    if isinstance(value, date):
        return value.strftime("%Y/%m/%d")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, float):
        # 整数として表せる値は小数点を落とす（500.0 → 500）
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def load_sheets(path: Path) -> dict[str, list[list[str]]]:
    """Excelファイルを読み、シート名 → 行データ の形にする。"""
    if not path.exists():
        raise WorkbookError(f"ファイルが見つかりません: {path}")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise WorkbookError(
            "openpyxl が入っていません。pip install -r requirements.txt を実行してください。"
        ) from exc

    try:
        book = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        raise WorkbookError(f"Excelファイルとして読めませんでした: {exc}") from exc

    result: dict[str, list[list[str]]] = {}
    try:
        for name in book.sheetnames:
            rows: list[list[str]] = []
            for row in book[name].iter_rows(values_only=True):
                rows.append([_cell_text(cell) for cell in row])
            # 末尾の空行を落とす（Excelは余分な空行を持つことがある）
            while rows and not any(rows[-1]):
                rows.pop()
            result[name] = rows
    finally:
        book.close()
    return result


def fetch_all(
    config: AppConfig, path: Path, genres: Iterable[str] | None = None
) -> dict[str, Any]:
    """設定に従って、Excelファイルから資料を組み立てる。

    戻り値の形は sheets.fetch_all と同じです。
    """
    by_sheet = load_sheets(path)
    target = set(genres) if genres is not None else None
    result: dict[str, dict[str, Any]] = {}

    for source in config.sources:
        if target is not None and source.genre not in target:
            continue
        bucket = result.setdefault(
            source.genre, {"docs": [], "note": "", "error": None, "notes": []}
        )
        rows = by_sheet.get(source.worksheet)
        if rows is None:
            bucket["notes"].append(f"{source.label}: シート「{source.worksheet}」が見つかりません")
            bucket["error"] = "; ".join(n for n in bucket["notes"] if "見つかりません" in n)
            continue

        if source.is_key_value:
            docs, notes = sheets._read_key_value(source, rows, config, None)
        else:
            docs, notes = sheets._read_table(source, rows, config, None, {})

        bucket["docs"].extend(docs)
        detail = f"{source.label} {len(docs)}件"
        if notes:
            detail += f"（{' / '.join(notes)}）"
        bucket["notes"].append(detail)

    for bucket in result.values():
        bucket["note"] = " / ".join(bucket["notes"])
    return result


def available_sheets(path: Path) -> list[str]:
    return list(load_sheets(path))
